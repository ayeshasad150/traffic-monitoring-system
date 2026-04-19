import ssl
ssl._create_default_https_context = ssl._create_unverified_context
import streamlit as st
import cv2
import numpy as np
from ultralytics import YOLO
import easyocr
import openpyxl
import os
import tempfile
from datetime import datetime

# ─── Page Config ───────────────────────────────────────────
st.set_page_config(
    page_title="AI Traffic Monitoring System",
    page_icon="🚦",
    layout="wide"
)

# ─── Custom CSS ────────────────────────────────────────────
st.markdown("""
<style>
    .main { background-color: #0e1117; }
    .stApp { background-color: #0e1117; }
    h1 { color: #00ff88 !important; }
    h2, h3 { color: #ffffff !important; }
    .violation-alert {
        background-color: #ff000033;
        border: 2px solid #ff0000;
        border-radius: 10px;
        padding: 10px;
        margin: 5px 0;
        color: #ff4444;
        font-weight: bold;
    }
    .success-box {
        background-color: #00ff8833;
        border: 2px solid #00ff88;
        border-radius: 10px;
        padding: 10px;
        color: #00ff88;
    }
    .metric-card {
        background: linear-gradient(135deg, #1a1a2e, #16213e);
        border-radius: 15px;
        padding: 20px;
        text-align: center;
        border: 1px solid #00ff8844;
    }
</style>
""", unsafe_allow_html=True)

# ─── Load Models ───────────────────────────────────────────
@st.cache_resource
def load_models():
    vehicle_model = YOLO("vehicle_model/weights/best.pt")
    helmet_model  = YOLO("helmet_model2/weights/best.pt")
    plate_model   = YOLO("plate_best.pt")
    reader        = easyocr.Reader(['en'], gpu=False)
    return vehicle_model, helmet_model, plate_model, reader

# ─── Excel Setup ───────────────────────────────────────────
def init_excel():
    path = "violations.xlsx"
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Violations"
        ws.append(["Time", "Violation Type", "Number Plate", "Vehicle Class"])
        wb.save(path)
    return path

def save_violation(violation_type, plate_text, vehicle_class="Unknown"):
    path = init_excel()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        violation_type,
        plate_text,
        vehicle_class
    ])
    wb.save(path)

# ─── Speed Estimation ──────────────────────────────────────
def estimate_speed(prev_center, curr_center, fps=30, ppm=8.8):
    if prev_center is None:
        return 0
    dx = curr_center[0] - prev_center[0]
    dy = curr_center[1] - prev_center[1]
    dist_px = np.sqrt(dx**2 + dy**2)
    dist_m  = dist_px / ppm
    speed   = dist_m * fps * 3.6
    return round(min(speed, 180), 1)

# ─── Process Frame ─────────────────────────────────────────
def process_frame(frame, vehicle_model, helmet_model,
                  plate_model, reader,
                  signal_line_y, prev_centers,
                  vehicle_count, violation_count,
                  conf_thresh=0.25):

    frame = cv2.resize(frame, (1280, 720))
    h, w = frame.shape[:2]
    annotated = frame.copy()

    # Signal line drawing
    cv2.line(annotated, (0, signal_line_y), (w, signal_line_y), (0, 0, 255), 3)
    cv2.putText(annotated, "STOP LINE",
                (10, signal_line_y - 10),
                cv2.FONT_HERSHEY_SIMPLEX, 0.8,
                (0, 0, 255), 2)

    # Vehicle Detection with tracking
    v_results = vehicle_model.track(
        frame, persist=True,
        conf=conf_thresh,
        iou=0.5,
        verbose=False
    )

    current_centers = {}
    violations_this_frame = []

    if v_results[0].boxes is not None and len(v_results[0].boxes) > 0:
        boxes = v_results[0].boxes

        for box in boxes:
            x1, y1, x2, y2 = map(int, box.xyxy[0])
            conf  = float(box.conf[0])
            cls   = int(box.cls[0])
            label = vehicle_model.names[cls]
            tid   = int(box.id[0]) if box.id is not None else -1

            # Skip if outside frame
            x1 = max(0, x1); y1 = max(0, y1)
            x2 = min(w, x2); y2 = min(h, y2)

            cx = (x1 + x2) // 2
            cy = (y1 + y2) // 2

            if tid >= 0:
                current_centers[tid] = (cx, cy)
                if tid not in prev_centers:
                    vehicle_count += 1

            # Speed estimation
            speed = 0
            if tid >= 0 and tid in prev_centers:
                speed = estimate_speed(prev_centers.get(tid), (cx, cy))

            color = (0, 255, 0)  # Green default

            # ── Signal Violation Detection ──
            is_violation = False
            if signal_line_y - 5 <= cy <= signal_line_y + 50:
                is_violation = True
                color = (0, 0, 255)  # Red
                violation_count += 1

                # Number Plate Detection
                plate_text = "Unknown"
                try:
                    roi = frame[y1:y2, x1:x2]
                    if roi.size > 0 and roi.shape[0] > 10 and roi.shape[1] > 10:
                        p_results = plate_model(roi, verbose=False, conf=0.3)
                        if p_results[0].boxes is not None and len(p_results[0].boxes) > 0:
                            pb = p_results[0].boxes[0]
                            px1, py1, px2, py2 = map(int, pb.xyxy[0])
                            px1 = max(0, px1); py1 = max(0, py1)
                            plate_roi = roi[py1:py2, px1:px2]
                            if plate_roi.size > 0:
                                ocr_result = reader.readtext(plate_roi)
                                if ocr_result:
                                    plate_text = ocr_result[0][1].upper()
                                    # Draw plate box
                                    cv2.rectangle(annotated,
                                                  (x1+px1, y1+py1),
                                                  (x1+px2, y1+py2),
                                                  (255, 255, 0), 2)
                                    cv2.putText(annotated, plate_text,
                                                (x1+px1, y1+py1-5),
                                                cv2.FONT_HERSHEY_SIMPLEX,
                                                0.5, (255, 255, 0), 2)
                except Exception:
                    pass

                save_violation("Signal Violation", plate_text, label)
                violations_this_frame.append(f"🚨 {label} - Plate: {plate_text}")

                # Violation alert on frame
                cv2.putText(annotated, "SIGNAL VIOLATION!",
                            (x1, y1 - 35),
                            cv2.FONT_HERSHEY_SIMPLEX,
                            0.7, (0, 0, 255), 2)

            # ── Helmet Detection (motorcycles) ──
            if label in ["motorbike", "motorcycle"]:
                try:
                    roi = frame[y1:y2, x1:x2]
                    if roi.size > 0 and roi.shape[0] > 20 and roi.shape[1] > 20:
                        h_results = helmet_model(roi, verbose=False, conf=0.3)
                        if h_results[0].boxes is not None:
                            for hbox in h_results[0].boxes:
                                hcls   = int(hbox.cls[0])
                                hlabel = helmet_model.names[hcls]
                                if "no" in hlabel.lower() or "No" in hlabel:
                                    color = (0, 165, 255)  # Orange
                                    cv2.putText(annotated, "NO HELMET!",
                                                (x1, y1 - 55),
                                                cv2.FONT_HERSHEY_SIMPLEX,
                                                0.7, (0, 165, 255), 2)
                                    save_violation("No Helmet", "N/A", label)
                                    violation_count += 1
                                    violations_this_frame.append(f"⛑️ No Helmet Detected!")
                except Exception:
                    pass

            # ── Draw Bounding Box ──
            cv2.rectangle(annotated, (x1, y1), (x2, y2), color, 2)

            # Label with speed
            display_label = f"{label}"
            if speed > 0:
                display_label += f" {speed}km/h"
            display_label += f" {conf:.0%}"

            # Background for text
            (tw, th), _ = cv2.getTextSize(
                display_label, cv2.FONT_HERSHEY_SIMPLEX, 0.55, 2)
            cv2.rectangle(annotated,
                          (x1, y1 - th - 8),
                          (x1 + tw + 4, y1),
                          color, -1)
            cv2.putText(annotated, display_label,
                        (x1 + 2, y1 - 5),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        0.55, (255, 255, 255), 2)

    # Update tracking
    prev_centers.update(current_centers)

    # ── Stats Overlay ──
    overlay = annotated.copy()
    cv2.rectangle(overlay, (0, 0), (300, 90), (0, 0, 0), -1)
    cv2.addWeighted(overlay, 0.6, annotated, 0.4, 0, annotated)

    cv2.putText(annotated, f"Vehicles: {vehicle_count}",
                (10, 35), cv2.FONT_HERSHEY_SIMPLEX,
                1.1, (0, 255, 136), 2)
    cv2.putText(annotated, f"Violations: {violation_count}",
                (10, 75), cv2.FONT_HERSHEY_SIMPLEX,
                1.1, (0, 0, 255), 2)

    return annotated, prev_centers, vehicle_count, violation_count, violations_this_frame

# ─── Main App ──────────────────────────────────────────────
def main():
    # Header
    st.markdown("# 🚦 AI Traffic Monitoring System")
    st.markdown("### Powered by YOLOv8 | Karachi Traffic Analysis")
    st.markdown("---")

    # Sidebar Settings
    st.sidebar.markdown("## ⚙️ Settings")
    signal_line = st.sidebar.slider("Signal Line Position (Y)", 50, 700, 300, 10)
    conf_thresh = st.sidebar.slider("Confidence Threshold", 0.1, 0.9, 0.25, 0.05)

    st.sidebar.markdown("---")
    st.sidebar.markdown("## 📊 Live Stats")
    vehicle_metric   = st.sidebar.empty()
    violation_metric = st.sidebar.empty()

    st.sidebar.markdown("---")
    st.sidebar.markdown("## 🚨 Recent Violations")
    violation_log = st.sidebar.empty()

    # Load Models
    with st.spinner("🔄 AI Models Load... Please Wait!"):
        try:
            vehicle_model, helmet_model, plate_model, reader = load_models()
            st.markdown('<div class="success-box">✅ All Models Successfully!</div>',
                        unsafe_allow_html=True)
        except Exception as e:
            st.error(f"❌ Model Load Error: {e}")
            st.stop()

    st.markdown("---")

    # Model Info
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown('<div class="metric-card">🚗<br><b>Vehicle Model</b><br>YOLOv8 | 95.6% mAP</div>',
                    unsafe_allow_html=True)
    with col2:
        st.markdown('<div class="metric-card">⛑️<br><b>Helmet Model</b><br>YOLOv8 | Helmet/No-Helmet</div>',
                    unsafe_allow_html=True)
    with col3:
        st.markdown('<div class="metric-card">🔤<br><b>Plate Model</b><br>YOLOv8 + EasyOCR</div>',
                    unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("## 📹 Traffic Video Upload Karo")

    video_file = st.file_uploader(
        "MP4, AVI, MOV format supported",
        type=["mp4", "avi", "mov", "mpeg4"]
    )

    if video_file is not None:
        # Save temp file
        tfile = tempfile.NamedTemporaryFile(delete=False, suffix='.mp4')
        tfile.write(video_file.read())
        tfile.close()

        cap = cv2.VideoCapture(tfile.name)

        if not cap.isOpened():
            st.error("❌ Video open nahi ho saka!")
            return

        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        fps = cap.get(cv2.CAP_PROP_FPS) or 30

        st.info(f"📹 Video: {total_frames} frames | {fps:.0f} FPS")

        # Progress bar
        progress_bar = st.progress(0)
        status_text  = st.empty()
        stframe      = st.empty()

        # State variables
        prev_centers    = {}
        vehicle_count   = 0
        violation_count = 0
        all_violations  = []
        frame_num       = 0

        # Process every 2nd frame for speed
        skip_frames = 2

        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break

            frame_num += 1

            # Skip frames for performance
            if frame_num % skip_frames != 0:
                continue

            # Process frame
            try:
                frame, prev_centers, vehicle_count, \
                    violation_count, violations_this_frame = process_frame(
                        frame, vehicle_model, helmet_model,
                        plate_model, reader,
                        signal_line, prev_centers,
                        vehicle_count, violation_count,
                        conf_thresh
                    )
            except Exception as e:
                continue

            # Collect violations
            if violations_this_frame:
                all_violations.extend(violations_this_frame)
                all_violations = all_violations[-10:]  # Keep last 10

            # Display frame
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            stframe.image(frame_rgb, use_container_width=True)

            # Update metrics
            vehicle_metric.metric("🚗 Total Vehicles", vehicle_count)
            violation_metric.metric("⚠️ Total Violations", violation_count)

            # Update violation log
            if all_violations:
                log_text = "\n\n".join(all_violations[-5:])
                violation_log.markdown(f"```\n{log_text}\n```")

            # Update progress
            progress = min(frame_num / max(total_frames, 1), 1.0)
            progress_bar.progress(progress)
            status_text.text(f"Processing: Frame {frame_num}/{total_frames}")

        cap.release()
        os.unlink(tfile.name)

        progress_bar.progress(1.0)
        status_text.text("✅ Processing Complete!")

        st.success(f"🎉 Video Processing Complete! | Vehicles: {vehicle_count} | Violations: {violation_count}")

        # Download Excel
        st.markdown("---")
        st.markdown("## 📥 Violations Report")

        if os.path.exists("violations.xlsx"):
            # Show table
            wb = openpyxl.load_workbook("violations.xlsx")
            ws = wb.active
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)

            if len(data) > 1:
                import pandas as pd
                df = pd.DataFrame(data[1:], columns=data[0])
                st.dataframe(df, use_container_width=True)

            # Download button
            with open("violations.xlsx", "rb") as f:
                st.download_button(
                    label="📥 Violations Excel Download",
                    data=f,
                    file_name="traffic_violations.xlsx",
                    mime="application/vnd.ms-excel"
                )
        else:
            st.info("ℹ️ No violation detect!")

if __name__ == "__main__":
    main()
    